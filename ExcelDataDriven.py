#open openpyxl in web
import openpyxl
Dict = {}
book = openpyxl.load_workbook("C:/Users/Selvakumar/Desktop/Pranesh/PythonDataDriverDemo.xlsx")
sheet = book.active
cell = sheet.cell(row=1, column=2)

print(cell.value)
print(sheet['B1'].value)
#above both syntax will print same value

#sheet.cell(row=1, column=2).value = "Rahul"
print(sheet.cell(row=1, column=2).value)
print(sheet.max_row)
print(sheet.max_column)

for i in range(1,sheet.max_row+1):
    if sheet.cell(row=i, column=1).value == "TC2":
        for j in range(2,sheet.max_column+1):
            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value  #loading excel data's into dictionary

print(Dict)



